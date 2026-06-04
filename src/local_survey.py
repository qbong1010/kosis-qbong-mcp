from __future__ import annotations

import hashlib
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
import yaml
from openpyxl import load_workbook

from .config import ROOT_DIR


SURVEY_YEAR = 2024
REFERENCE_YEAR = 2023
LOCAL_REGION_CODE = "32030"
LOCAL_REGION_NAME = "강릉시"
LOCAL_PROVINCE_NAME = "강원특별자치도"


@dataclass(frozen=True)
class WorkbookLayout:
    table_no: int
    table_name: str
    industry_level_hint: str
    breakdown_type: str
    name_col: int
    code_cols: tuple[int, ...]
    area_header_row: int
    metric_header_row: int | None
    data_start_row: int
    value_start_col: int
    paired_metrics: bool
    fixed_indicator: str | None = None


LAYOUTS: dict[int, WorkbookLayout] = {
    2: WorkbookLayout(2, "산업소분류 및 읍면동별 사업체수, 종사자수", "소분류", "none", 3, (0, 1, 2), 3, 4, 5, 4, True),
    3: WorkbookLayout(3, "산업중분류, 종사자규모 및 읍면동별 사업체수, 종사자수", "중분류", "worker_size", 2, (0, 1), 3, 4, 5, 3, True),
    4: WorkbookLayout(4, "산업중분류, 사업체구분 및 읍면동별 사업체수, 종사자수", "중분류", "establishment_type", 0, (), 3, 4, 5, 1, True),
    5: WorkbookLayout(5, "산업중분류, 조직형태 및 읍면동별 사업체수, 종사자수", "중분류", "organization_type", 0, (), 3, 4, 5, 1, True),
    6: WorkbookLayout(6, "산업중분류, 종사상지위 및 읍면동별 종사자수", "중분류", "worker_status", 0, (), 3, None, 4, 1, False, "employees"),
    7: WorkbookLayout(7, "산업중분류, 대표자남여 및 읍면동별 사업체수", "중분류", "representative_gender", 0, (), 3, None, 4, 1, False, "establishments"),
    8: WorkbookLayout(8, "산업중분류, 종사자남여 및 읍면동별 종사자수", "중분류", "worker_gender", 0, (), 3, None, 4, 1, False, "employees"),
    9: WorkbookLayout(9, "산업중분류, 대표자연령대 및 읍면동별 사업체수, 종사자수", "중분류", "representative_age", 0, (), 3, 4, 5, 1, True),
}


def local_source_root() -> Path:
    return ROOT_DIR / "data" / "sources" / "local_gov" / "gangneung" / "business_survey" / str(SURVEY_YEAR)


def source_workbook_paths() -> list[Path]:
    source_dirs = [p for p in ROOT_DIR.iterdir() if p.is_dir() and "사업체조사" in p.name and "강릉" in p.name]
    if not source_dirs:
        return []
    return sorted(source_dirs[0].glob("*.xlsx"))


def archive_raw_workbooks(paths: list[Path]) -> list[Path]:
    raw_dir = local_source_root() / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    archived: list[Path] = []
    for path in paths:
        target = raw_dir / path.name
        if not target.exists() or sha256_file(target) != sha256_file(path):
            shutil.copy2(path, target)
        archived.append(target)
    return archived


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_manifest(paths: list[Path]) -> tuple[Path, Path]:
    root = local_source_root()
    checksums = [
        {
            "file_name": path.name,
            "relative_path": str(path.relative_to(root)).replace("\\", "/"),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        for path in paths
    ]
    manifest = {
        "source_type": "local_gov",
        "source_org": LOCAL_REGION_NAME,
        "province": LOCAL_PROVINCE_NAME,
        "survey_name": "강릉시 사업체조사 결과",
        "survey_year": SURVEY_YEAR,
        "reference_year": REFERENCE_YEAR,
        "region_code": LOCAL_REGION_CODE,
        "region_name": LOCAL_REGION_NAME,
        "raw_files": checksums,
        "storage_policy": "raw xlsx is immutable; analysis uses normalized DuckDB/Parquet tables",
    }
    manifest_path = root / "manifest.yaml"
    checksum_path = root / "checksum.json"
    manifest_path.write_text(yaml.safe_dump(manifest, allow_unicode=True, sort_keys=False), encoding="utf-8")
    checksum_path.write_text(json.dumps(checksums, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path, checksum_path


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _parse_value(value: Any) -> tuple[float | None, str]:
    text = _clean(value).replace(",", "")
    if text == "":
        return None, "missing"
    if text in {"*", "-", "X", "x", "…", "..."}:
        return None, "suppressed"
    try:
        return float(text), "valid"
    except ValueError:
        return None, "missing"


def _table_no(path: Path) -> int | None:
    match = re.match(r"^(\d+)", path.name)
    return int(match.group(1)) if match else None


def _normalize_code(raw_code: str, name: str, parent_major: str | None = None) -> str:
    code = _clean(raw_code)
    text = _clean(name)
    if text.startswith("TT") or code == "TT":
        return "TT"
    if not code:
        match = re.match(r"^([A-Z]|\d{2,5})\.", text)
        if match:
            code = match.group(1)
    if code in {"**", ""}:
        return parent_major or ""
    if re.fullmatch(r"\d{2,5}", code):
        if code[:2].isdigit() and 10 <= int(code[:2]) <= 34:
            return f"C{code}"
        return code
    return code


def _clean_industry_name(text: str) -> str:
    text = _clean(text)
    text = re.sub(r"^(TT|[A-Z]|\d{2,5})\.\s*", "", text)
    return text


def _industry_level(code: str) -> str:
    if code == "TT":
        return "전체"
    if re.fullmatch(r"[A-Z]", code):
        return "대분류"
    if re.fullmatch(r"[A-Z]\d{2}", code) or re.fullmatch(r"\d{2}", code):
        return "중분류"
    if re.fullmatch(r"[A-Z]\d{3}", code) or re.fullmatch(r"\d{3}", code):
        return "소분류"
    if re.fullmatch(r"[A-Z]\d{4}", code) or re.fullmatch(r"\d{4}", code):
        return "세분류"
    if re.fullmatch(r"[A-Z]\d{5}", code) or re.fullmatch(r"\d{5}", code):
        return "세세분류"
    return "unknown"


def _is_industry_row(layout: WorkbookLayout, row: tuple[Any, ...]) -> bool:
    text = _clean(row[layout.name_col] if layout.name_col < len(row) else "")
    if not text:
        return False
    if layout.code_cols and any(_clean(row[idx] if idx < len(row) else "") not in {"", "**"} for idx in layout.code_cols):
        return True
    return bool(re.match(r"^(TT|[A-Z]|\d{2,5})\.", text))


def _industry_from_row(layout: WorkbookLayout, row: tuple[Any, ...], current_major: str | None) -> tuple[str, str, str, str | None]:
    name_text = _clean(row[layout.name_col])
    raw_code = ""
    for idx in reversed(layout.code_cols):
        value = _clean(row[idx] if idx < len(row) else "")
        if value and value != "**":
            raw_code = value
            break
    code = _normalize_code(raw_code, name_text, current_major)
    if re.fullmatch(r"[A-Z]", code):
        current_major = code
    name = _clean_industry_name(name_text)
    return code, name, _industry_level(code), current_major


def _build_value_columns(ws: Any, layout: WorkbookLayout) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(min_row=1, max_row=max(layout.area_header_row, layout.metric_header_row or 0), values_only=True))
    area_row = rows[layout.area_header_row - 1]
    metric_row = rows[layout.metric_header_row - 1] if layout.metric_header_row else None
    columns: list[dict[str, Any]] = []
    current_area = ""
    col = layout.value_start_col
    while col < ws.max_column:
        area = _clean(area_row[col] if col < len(area_row) else "")
        if area:
            current_area = area
        if not current_area:
            col += 1
            continue
        if layout.paired_metrics:
            metric_text = _clean(metric_row[col] if metric_row and col < len(metric_row) else "")
            if "사업체" in metric_text:
                indicator = "establishments"
                unit = "개"
            elif "종사자" in metric_text:
                indicator = "employees"
                unit = "명"
            else:
                col += 1
                continue
        else:
            indicator = layout.fixed_indicator or "value"
            unit = "명" if indicator == "employees" else "개"
        columns.append({"col": col, "admin_area_name": current_area, "indicator": indicator, "unit": unit})
        col += 1
    return columns


def extract_raw_cells(paths: list[Path]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for path in paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    records.append(
                        {
                            "source_file": path.name,
                            "sheet_name": sheet,
                            "row_index": cell.row,
                            "column_index": cell.column,
                            "cell_value": str(cell.value),
                        }
                    )
    return pd.DataFrame.from_records(records)


def parse_observations(paths: list[Path]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for path in paths:
        table_no = _table_no(path)
        if table_no == 1:
            records.extend(_parse_total_workbook(path))
            continue
        if table_no not in LAYOUTS:
            continue
        layout = LAYOUTS[table_no]
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        value_columns = _build_value_columns(ws, layout)
        current_industry: dict[str, str] | None = None
        current_major: str | None = None

        for excel_row_index, row in enumerate(ws.iter_rows(min_row=layout.data_start_row, values_only=True), start=layout.data_start_row):
            label = _clean(row[layout.name_col] if layout.name_col < len(row) else "")
            if not label:
                continue
            if _is_industry_row(layout, row):
                code, name, level, current_major = _industry_from_row(layout, row, current_major)
                current_industry = {"industry_code": code, "industry_name": name, "industry_level": level}
                breakdown_value = "전체"
            elif current_industry is not None:
                breakdown_value = label
            else:
                continue

            for col in value_columns:
                raw_value = row[col["col"]] if col["col"] < len(row) else None
                value, quality_flag = _parse_value(raw_value)
                records.append(
                    {
                        "source_type": "local_gov",
                        "source_org": LOCAL_REGION_NAME,
                        "source_file": path.name,
                        "table_no": table_no,
                        "table_name": layout.table_name,
                        "sheet_name": ws.title,
                        "excel_row": excel_row_index,
                        "survey_year": SURVEY_YEAR,
                        "reference_year": REFERENCE_YEAR,
                        "region_code": LOCAL_REGION_CODE,
                        "region_name": LOCAL_REGION_NAME,
                        "admin_area_name": col["admin_area_name"],
                        "industry_code": current_industry["industry_code"],
                        "industry_name": current_industry["industry_name"],
                        "industry_level": current_industry["industry_level"],
                        "breakdown_type": layout.breakdown_type,
                        "breakdown_value": breakdown_value,
                        "indicator": col["indicator"],
                        "unit": col["unit"],
                        "value": value,
                        "quality_flag": quality_flag,
                        "raw_value": _clean(raw_value),
                    }
                )
    return pd.DataFrame.from_records(records)


def _parse_total_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[dict[str, Any]] = []
    value_specs = [
        {"col": 6, "indicator": "establishments", "unit": "개", "breakdown_type": "none", "breakdown_value": "전체"},
        {"col": 7, "indicator": "employees", "unit": "명", "breakdown_type": "none", "breakdown_value": "전체"},
        {"col": 8, "indicator": "employees", "unit": "명", "breakdown_type": "worker_gender", "breakdown_value": "남"},
        {"col": 9, "indicator": "employees", "unit": "명", "breakdown_type": "worker_gender", "breakdown_value": "여"},
    ]
    current_major: str | None = None
    for excel_row_index, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        name_text = _clean(row[5] if len(row) > 5 else "")
        if not name_text:
            continue
        raw_code = ""
        for idx in range(4, -1, -1):
            value = _clean(row[idx] if len(row) > idx else "")
            if value and value != "**":
                raw_code = value
                break
        code = _normalize_code(raw_code, name_text, current_major)
        if re.fullmatch(r"[A-Z]", code):
            current_major = code
        industry = {"industry_code": code, "industry_name": _clean_industry_name(name_text), "industry_level": _industry_level(code)}
        for spec in value_specs:
            raw_value = row[spec["col"]] if len(row) > spec["col"] else None
            value, quality_flag = _parse_value(raw_value)
            records.append(
                {
                    "source_type": "local_gov",
                    "source_org": LOCAL_REGION_NAME,
                    "source_file": path.name,
                    "table_no": 1,
                    "table_name": "산업세세분류별 총괄",
                    "sheet_name": ws.title,
                    "excel_row": excel_row_index,
                    "survey_year": SURVEY_YEAR,
                    "reference_year": REFERENCE_YEAR,
                    "region_code": LOCAL_REGION_CODE,
                    "region_name": LOCAL_REGION_NAME,
                    "admin_area_name": LOCAL_REGION_NAME,
                    "industry_code": industry["industry_code"],
                    "industry_name": industry["industry_name"],
                    "industry_level": industry["industry_level"],
                    "breakdown_type": spec["breakdown_type"],
                    "breakdown_value": spec["breakdown_value"],
                    "indicator": spec["indicator"],
                    "unit": spec["unit"],
                    "value": value,
                    "quality_flag": quality_flag,
                    "raw_value": _clean(raw_value),
                }
            )
    return records


def write_local_store(observations: pd.DataFrame, raw_cells: pd.DataFrame) -> dict[str, str | int]:
    processed_dir = ROOT_DIR / "data" / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    db_path = processed_dir / "local_business_survey.duckdb"
    parquet_path = processed_dir / "gangneung_business_survey_2024.parquet"
    raw_cells_parquet_path = processed_dir / "gangneung_business_survey_2024_raw_cells.parquet"

    with duckdb.connect(str(db_path)) as con:
        con.register("observations_df", observations)
        con.execute("CREATE OR REPLACE TABLE business_survey_observations AS SELECT * FROM observations_df")
        con.register("raw_cells_df", raw_cells)
        con.execute("CREATE OR REPLACE TABLE business_survey_raw_cells AS SELECT * FROM raw_cells_df")
        con.execute(
            """
            CREATE OR REPLACE VIEW gangneung_industry_middle_totals AS
            SELECT
              reference_year,
              region_code,
              region_name,
              admin_area_name,
              industry_code,
              industry_name,
              indicator,
              unit,
              value,
              quality_flag,
              source_file
            FROM business_survey_observations
            WHERE table_no = 3
              AND industry_level = '중분류'
              AND breakdown_value = '전체'
              AND admin_area_name = '강릉시'
              AND indicator IN ('establishments', 'employees')
            """
        )
        con.execute(
            """
            CREATE OR REPLACE VIEW gangneung_manufacturing_middle_totals AS
            SELECT *
            FROM gangneung_industry_middle_totals
            WHERE industry_code LIKE 'C__'
            """
        )
        con.execute(f"COPY business_survey_observations TO '{parquet_path.as_posix()}' (FORMAT PARQUET)")
        con.execute(f"COPY business_survey_raw_cells TO '{raw_cells_parquet_path.as_posix()}' (FORMAT PARQUET)")

    observations.to_csv(processed_dir / "gangneung_business_survey_2024_observations.csv", index=False, encoding="utf-8-sig")
    return {
        "duckdb_path": str(db_path),
        "observations_parquet_path": str(parquet_path),
        "raw_cells_parquet_path": str(raw_cells_parquet_path),
        "observation_count": int(len(observations)),
        "raw_cell_count": int(len(raw_cells)),
    }


def ingest_gangneung_business_survey() -> dict[str, Any]:
    source_paths = source_workbook_paths()
    if not source_paths:
        raise FileNotFoundError("강릉시 사업체조사 엑셀 원자료 폴더를 찾지 못했습니다.")
    archived_paths = archive_raw_workbooks(source_paths)
    manifest_path, checksum_path = write_manifest(archived_paths)
    observations = parse_observations(archived_paths)
    raw_cells = extract_raw_cells(archived_paths)
    outputs = write_local_store(observations, raw_cells)
    return {
        "source_files": len(archived_paths),
        "manifest_path": str(manifest_path),
        "checksum_path": str(checksum_path),
        **outputs,
    }


def query_local_business_survey(
    *,
    industry_level: str | None = None,
    indicator: str | None = None,
    admin_area_name: str | None = None,
    breakdown_type: str | None = None,
    breakdown_value: str | None = None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    db_path = ROOT_DIR / "data" / "processed" / "local_business_survey.duckdb"
    if not db_path.exists():
        raise FileNotFoundError("local_business_survey.duckdb가 없습니다. 먼저 ingestion을 실행하세요.")
    clauses: list[str] = []
    params: list[Any] = []
    filters = {
        "industry_level": industry_level,
        "indicator": indicator,
        "admin_area_name": admin_area_name,
        "breakdown_type": breakdown_type,
        "breakdown_value": breakdown_value,
    }
    for column, value in filters.items():
        if value is not None:
            clauses.append(f"{column} = ?")
            params.append(value)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    sql = f"""
        SELECT *
        FROM business_survey_observations
        {where}
        ORDER BY table_no, excel_row, admin_area_name, indicator
        LIMIT ?
    """
    params.append(limit)
    with duckdb.connect(str(db_path), read_only=True) as con:
        return con.execute(sql, params).fetchdf().to_dict(orient="records")


def fetch_local_industry_middle_totals(manufacturing_only: bool = True) -> list[dict[str, Any]]:
    db_path = ROOT_DIR / "data" / "processed" / "local_business_survey.duckdb"
    if not db_path.exists():
        raise FileNotFoundError("local_business_survey.duckdb가 없습니다. 먼저 ingestion을 실행하세요.")
    view_name = "gangneung_manufacturing_middle_totals" if manufacturing_only else "gangneung_industry_middle_totals"
    with duckdb.connect(str(db_path), read_only=True) as con:
        return con.execute(f"SELECT * FROM {view_name} ORDER BY industry_code, indicator").fetchdf().to_dict(orient="records")
